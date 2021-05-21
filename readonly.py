from openpyxl import load_workbook
import pymysql
import re

def strprocess(str):
    cop = re.compile("[^\u4e00-\u9fa5^a-z^A-Z^0-9]")  # 匹配不是中文、大小写、数字的其他字符
    return cop.sub('', str)  # 将string1中匹配到的字符替换成空字符

con = pymysql.connect(user="inventory", passwd="123456", db="inventory", host="127.0.0.1", local_infile=1)
cur = con.cursor()
cur.execute("set names utf8")
cur.execute("SET character_set_connection=utf8;")

excel_path = r'E:\564194E_CCM_FulfillReport_53658764_20201109003456404.xlsx'
wb = load_workbook(filename=excel_path, read_only=True)
for ws_name in wb.sheetnames:
    wb[ws_name].reset_dimensions()
    table_name = strprocess(ws_name)
    drop_sql = "drop table if exists {}".format(table_name)
    cur.execute(drop_sql)
    create_sql = "create table {}(".format(table_name)
    title_list = []
    for row in wb[ws_name].rows:
        if len(title_list) > 0:
            data_list = []
            for cell in row:
                if cell:
                    data_list.append(str(cell.value)[:85])
                else:
                    data_list.append("")
            insert_sql = "insert into {} values({})".format(table_name, str(data_list[:len(title_list)]).replace('[', '').replace(']', ''))
            cur.execute(insert_sql)
        else:
            for cell in row:
                if cell.value:
                    title_list.append(strprocess(cell.value))
            for title in title_list[:-1]:
                create_sql += "{} varchar(255),".format(title)
            create_sql += "{} varchar(255))".format(title_list[-1])
            cur.execute(create_sql)
con.commit()
con.close()
wb.close()
