import csv
from openpyxl import load_workbook, Workbook

site_mapping_dict = {}
wb = load_workbook('Inventory Selection List.xlsx')
site_mapping_st = wb['Site Mapping']
for i in range(1, site_mapping_st.max_row):
    site_mapping_dict.update({site_mapping_st.cell(i + 1, 2).value + '@' + site_mapping_st.cell(i + 1, 3).value: site_mapping_st.cell(i + 1, 4).value})

d1 = {}
with open('Inventory_Subrack_20210513_122725.csv', 'r', encoding='utf-8') as csv_file:
    data_list = csv.DictReader(csv_file)
    for data_dict in data_list:
        netype = data_dict.get('NEType')
        nefdn = data_dict.get('NEFdn')
        nename = data_dict.get('NEName')
        frame_type = data_dict.get('Frame Type')
        sn = data_dict.get('SN(Bar Code)')
        if netype in ['MBTS']:
            continue
        if frame_type not in ['BBU3900', 'BBU5900', 'BBU3910']:
            continue
        if sn not in [None, '', ' ']:
            continue
        temp1 = d1.get(sn + '@' + frame_type)
        if temp1:
            temp1.add(nefdn + '@' + nename)
            d1.update({sn + '@' + frame_type: temp1})
        else:
            d1.update({sn + '@' + frame_type: {nefdn + '@' + nename}})
wb = Workbook()
ws = wb.active
ws.title = 'bbu'
start_row = 1
for sn, nefdn_nename_set in d1.items():
    ws.cell(start_row, 1, sn)
    start_col = 2
    for nefdn_nename in nefdn_nename_set:
        ws.cell(start_row, start_col, nefdn_nename)
        start_col += 1
    start_row += 1
wb.save('bbu.xlsx')
