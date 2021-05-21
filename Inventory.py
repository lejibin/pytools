import csv
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


class InventoryReportTool:
    def __init__(self, work_path=None):
        if work_path:
            self.work_path = work_path
        else:
            self.work_path = os.getcwd()
        self.inventory_file_list = self.get_inventory_files(self.work_path)
        self.result_item_dict = {}
        self.result_category_dict = {}
        self.title_dict = {}
        self.bom_manufacturer_dict = {}
        self.select_dict = {'select_board': {}, 'select_subrack': {}, 'select_cabinet': {}}
        self.category_dict = {}
        self.bom_dict = {}
        self.site_mapping_dict = {}
        self.selection()
        self.data_load()
        self.data_output()

    def get_inventory_files(self, root_path, all_files=None):
        if all_files is None:
            all_files = []
        files = os.listdir(root_path)
        for file in files:
            if not os.path.isdir(root_path + '/' + file):  # not a dir
                if file.endswith('.csv') or file.endswith('.CSV'):
                    all_files.append(root_path + '/' + file)
            else:  # is a dir
                self.get_inventory_files((root_path + '/' + file), all_files)
        return all_files

    def selection(self):
        wb = load_workbook(os.path.join(self.work_path, 'Inventory Selection List.xlsx'))
        board_st = wb['Board']
        subrack_st = wb['Subrack']
        cabinet_st = wb['Cabinet']
        category_st = wb['Category']
        site_mapping_st = wb['Site Mapping']
        select_board_dict = {}
        for i in range(1, board_st.max_row):
            select_board_dict.update({board_st.cell(i + 1, 1).value: board_st.cell(i + 1, 2).value})
        self.select_dict.update({'select_board': select_board_dict})
        select_subrack_dict = {}
        for i in range(1, subrack_st.max_row):
            select_subrack_dict.update({subrack_st.cell(i + 1, 1).value: subrack_st.cell(i + 1, 2).value})
        self.select_dict.update({'select_subrack': select_subrack_dict})
        select_cabinet_dict = {}
        for i in range(1, cabinet_st.max_row):
            select_cabinet_dict.update({cabinet_st.cell(i + 1, 1).value: cabinet_st.cell(i + 1, 2).value})
        self.select_dict.update({'select_cabinet': select_cabinet_dict})
        for i in range(1, category_st.max_row):
            category = category_st.cell(i + 1, 1).value + '_' + (category_st.cell(i + 1, 2).value if category_st.cell(i + 1, 2).value else '') + '_' + (category_st.cell(i + 1, 3).value if category_st.cell(i + 1, 3).value else '')
            bom_code = category_st.cell(i + 1, 4).value
            self.category_dict.update({bom_code: category})
            if category in self.bom_dict.keys():
                bom_code_str = self.bom_dict.get(category) + '/' + bom_code
            else:
                bom_code_str = bom_code
            self.bom_dict.update({category: bom_code_str})
        for i in range(1, site_mapping_st.max_row):
            # self.site_mapping_dict.update({site_mapping_st.cell(i + 1, 2).value + '@' + site_mapping_st.cell(i + 1, 3).value: site_mapping_st.cell(i + 1, 4).value})
            self.site_mapping_dict.update({site_mapping_st.cell(i + 1, 2).value + '@' + site_mapping_st.cell(i + 1, 3).value: site_mapping_st.cell(i + 1, 5).value + '@' + site_mapping_st.cell(i + 1, 6).value + '@' + site_mapping_st.cell(i + 1, 7).value})

    def format_site_name(self, site_name):
        # site_name = site_name.upper()
        # if '_' in site_name:
        #     return site_name[1:4] + '_' + site_name[-2:]
        # else:
        #     return site_name
        if site_name not in self.site_mapping_dict.keys():
            print('Unknown Site: ' + site_name)
            return site_name
        s = self.site_mapping_dict.get(site_name)
        if s not in [None, '', ' ']:
            return s
        else:
            return site_name

    def data_load(self):
        sn_dict = {}
        for inventory_file in self.inventory_file_list:
            if ('BOARD' not in inventory_file.upper()) and ('SUBRACK' not in inventory_file.upper()) and (
                    'CABINET' not in inventory_file.upper()):
                # 打印未识别的文件名
                print('Unknown File: ' + inventory_file)
                continue
            with open(inventory_file, 'r', encoding='utf-8') as csv_file:
                data = csv.DictReader(csv_file)
                if 'BOARD' in inventory_file.upper():
                    select_key = 'Board Name'
                    select_dict = self.select_dict.get('select_board')
                elif 'SUBRACK' in inventory_file.upper():
                    select_key = 'Frame Type'
                    select_dict = self.select_dict.get('select_subrack')
                elif 'CABINET' in inventory_file.upper():
                    select_key = 'Rack Type'
                    select_dict = self.select_dict.get('select_cabinet')
                try:
                    for row in data:
                        select_type = row.get(select_key)
                        bom_code = row.get('PN(BOM Code/Item)')
                        manufacturer_data = row.get('Manufacturer Data')
                        sn = row.get('SN(Bar Code)')
                        netype = row.get('NEType')
                        if 'MBTS' == netype:
                            continue
                        # 判断类型取值是否为空
                        if select_type in ['', ' ', None]:
                            continue
                        # 判断未识别的类型名并打印
                        if select_type not in select_dict.keys():
                            print('Unknown Type: ' + inventory_file + '|' + select_type)
                            continue
                        # 判断类型是否需要统计
                        if select_dict.get(select_type) not in ['1', 1]:
                            continue
                        # 判断BOM Code取值是否为空
                        if bom_code in ['', ' ', None]:
                            continue
                        # 判断SN是否已经统计过
                        if sn in sn_dict.keys():
                            continue
                        site_name = self.format_site_name(row.get('NEFdn') + '@' + row.get('NEName'))
                        sn_dict.update({sn: site_name})
                        self.bom_manufacturer_dict.update({bom_code: manufacturer_data})
                        if bom_code not in self.category_dict.keys():
                            # 打印未识别的BOM Code
                            print('Unknown BOM: ' + inventory_file + '|' + select_type + '|' + bom_code)
                            category = 'Unknown'
                        else:
                            category = self.category_dict.get(bom_code)
                        # 更新输出总清单
                        if category in self.title_dict.keys():
                            bom_code_set = self.title_dict.get(category)
                            bom_code_set.add(bom_code)
                        else:
                            bom_code_set = {bom_code}
                        self.title_dict.update({category: bom_code_set})
                        # 如果结果中没有站点数据
                        if site_name not in self.result_item_dict.keys():
                            self.result_item_dict.update({site_name: {bom_code: 1}})
                        # 如果结果中有站点数据
                        else:
                            v1 = self.result_item_dict.get(site_name)
                            if bom_code not in v1.keys():
                                v1.update({bom_code: 1})
                                self.result_item_dict.update({site_name: v1})
                            else:
                                count = v1.get(bom_code) + 1
                                v1.update({bom_code: count})
                                self.result_item_dict.update({site_name: v1})
                        # 如果结果中没有类型数据
                        if category not in self.result_category_dict.keys():
                            self.result_category_dict.update({category: 1})
                        # 如果结果中有类型数据
                        else:
                            c = self.result_category_dict.get(category) + 1
                            self.result_category_dict.update({category: c})
                except:
                    print('Coding Err: ' + inventory_file + '|' + row.get('NEFdn'))

    def data_output(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Detail'
        col_n = 2
        title_list = sorted(self.title_dict.items())
        for title in title_list:
            category = title[0]
            bom_code_set = title[1]
            ws.cell(1, col_n, category)
            for bom_code in bom_code_set:
                ws.cell(2, col_n, bom_code)
                ws.cell(3, col_n, self.bom_manufacturer_dict.get(bom_code))
                row_n = 4
                for site_name, v in self.result_item_dict.items():
                    ws.cell(row_n, 1, site_name)
                    ws.cell(row_n, col_n, v.get(bom_code))
                    row_n = row_n + 1
                col_n = col_n + 1
        wb.create_sheet('Summary', 0)
        ws1 = wb['Summary']
        ws1.cell(1, 1, 'Category')
        ws1.cell(1, 2, 'Band')
        ws1.cell(1, 3, '5G Ready')
        ws1.cell(1, 4, 'BOM Code')
        ws1.cell(1, 5, 'Total')
        r = 2
        for category_str, total in self.result_category_dict.items():
            if 'Ignore' == category_str:
                continue
            bom_code_str = self.bom_dict.get(category_str)
            category_arr = category_str.split('_')
            ws1.cell(r, 1, category_arr[0])
            ws1.cell(r, 2, category_arr[1])
            ws1.cell(r, 3, category_arr[2])
            ws1.cell(r, 4, bom_code_str)
            ws1.cell(r, 5, total)
            r = r + 1
        file_date = datetime.datetime.now().strftime('%Y%m%d %H%M%S')
        wb.save(os.path.join(self.work_path, 'Inventory Summary ' + file_date + '.xlsx'))


if __name__ == '__main__':
    inventoryReportTool = InventoryReportTool(r'D:\VIVO\3. 存量')
